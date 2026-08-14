# -*- coding: utf-8 -*-
"""Pre-release validation for deployable modeling artifacts."""

from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import joblib


@dataclass
class ReleaseCheck:
    name: str
    passed: bool
    detail: str


@dataclass
class ReleaseValidationResult:
    passed: bool
    checks: List[ReleaseCheck]

    def as_dict(self) -> Dict[str, Any]:
        return {
            "passed": self.passed,
            "checks": [asdict(check) for check in self.checks],
        }


def _load_artifact(source: Union[str, Path, Dict[str, Any]]) -> Dict[str, Any]:
    if isinstance(source, dict):
        return source
    path = Path(source)
    if path.is_dir():
        path = path / "scoring_artifact.pkl"
    artifact = joblib.load(path)
    if isinstance(artifact, dict) and "scoring_artifact" in artifact:
        artifact = artifact["scoring_artifact"]
    if not isinstance(artifact, dict):
        raise ValueError("Scoring artifact must be a dictionary")
    return artifact


def validate_release(
    source: Union[str, Path, Dict[str, Any]],
    *,
    report_path: Optional[Union[str, Path]] = None,
    metrics: Optional[Dict[str, Any]] = None,
    min_auc: Optional[float] = None,
    max_psi: Optional[float] = None,
) -> ReleaseValidationResult:
    """Validate artifact completeness, report contract, and metric gates."""
    checks: List[ReleaseCheck] = []
    try:
        artifact = _load_artifact(source)
        checks.append(ReleaseCheck("artifact_loadable", True, "artifact loaded"))
    except Exception as exc:
        checks.append(ReleaseCheck("artifact_loadable", False, str(exc)))
        return ReleaseValidationResult(False, checks)

    required = ["artifact_version", "task", "feature_columns", "model", "preprocessor"]
    if artifact.get("task", "classification") == "classification":
        required.extend(["binner", "selector", "woe_feature_columns", "selected_features"])
    missing = [key for key in required if key not in artifact]
    checks.append(
        ReleaseCheck(
            "artifact_contract",
            not missing,
            "all required keys present" if not missing else f"missing keys: {missing}",
        )
    )

    features = artifact.get("feature_columns", [])
    selected = artifact.get("selected_features", [])
    checks.append(
        ReleaseCheck(
            "feature_contract",
            bool(features) and bool(selected),
            f"drivers={len(features)}, selected={len(selected)}",
        )
    )

    if report_path is None and not isinstance(source, dict):
        candidate = Path(source)
        if candidate.is_dir():
            reports = sorted(candidate.glob("Model_Report_*.xlsx"))
            report_path = reports[-1] if reports else candidate / "Model_Report_1.xlsx"
        else:
            report_path = candidate.parent / "Model_Report_1.xlsx"
    if report_path is not None:
        report = Path(report_path)
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(report, read_only=True)
            required_sheets = {"Overview_Performance", "Artifact_Metadata"}
            missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
            checks.append(
                ReleaseCheck(
                    "report_contract",
                    not missing_sheets,
                    "core report sheets present"
                    if not missing_sheets
                    else f"missing sheets: {missing_sheets}",
                )
            )
        except Exception as exc:
            checks.append(ReleaseCheck("report_contract", False, str(exc)))

    metrics = metrics or artifact.get("metadata", {}).get("metrics", {})
    if min_auc is not None:
        auc = metrics.get("auc_roc")
        checks.append(
            ReleaseCheck(
                "min_auc",
                auc is not None and float(auc) >= min_auc,
                f"auc_roc={auc}, required>={min_auc}",
            )
        )
    if max_psi is not None:
        psi = metrics.get("score_psi", metrics.get("psi"))
        checks.append(
            ReleaseCheck(
                "max_psi",
                psi is not None and float(psi) <= max_psi,
                f"psi={psi}, required<={max_psi}",
            )
        )

    return ReleaseValidationResult(
        passed=all(check.passed for check in checks),
        checks=checks,
    )
