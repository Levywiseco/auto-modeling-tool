"""Deterministic Excel audit reports for training and independent evaluation."""

from pathlib import Path
from typing import Any, Optional, Union

import polars as pl

from ..core.logger import logger


def _rows(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, pl.DataFrame):
        return value.to_dicts()
    if isinstance(value, dict):
        if value and all(isinstance(item, dict) for item in value.values()):
            rows = []
            for key, item in value.items():
                row = {"key": key}
                row.update(item)
                rows.append(row)
            return rows
        return [{"key": key, "value": value_item} for key, value_item in value.items()]
    if isinstance(value, list):
        return value if all(isinstance(item, dict) for item in value) else [
            {"value": item} for item in value
        ]
    return [{"value": value}]


def _json_safe(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    return str(value)


def _append_table(sheet: Any, value: Any, *, header_font: Any) -> None:
    rows = _rows(value)
    if not rows:
        sheet.append(["No records"])
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
    for row in rows:
        sheet.append([_json_safe(row.get(header)) for header in headers])


def write_model_report(
    output_dir: Union[str, Path],
    metrics: dict[str, Any],
    *,
    feature_importance: Optional[pl.DataFrame] = None,
    metadata: Optional[dict[str, Any]] = None,
    tables: Optional[dict[str, Any]] = None,
    filename: Optional[str] = None,
) -> Optional[Path]:
    """Write a stable, multi-sheet Model Report workbook.

    The core sheets are always present. Additional audit tables are supplied
    through the tables argument so pipelines can add binning, PSI, segment,
    temporal, benchmark, SHAP, or refinement diagnostics without changing
    the writer.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.warning("openpyxl is not installed; skipping Excel report")
        return None

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    report_path = output / filename if filename else None
    if report_path is None:
        index = 1
        while (output / f"Model_Report_{index}.xlsx").exists():
            index += 1
        report_path = output / f"Model_Report_{index}.xlsx"

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    overview = workbook.active
    overview.title = "Overview_Performance"
    overview.append(["Metric", "Value"])
    for cell in overview[1]:
        cell.font = header_font
        cell.fill = header_fill
    for key, value in (metrics or {}).items():
        overview.append([key, _json_safe(value)])

    index_sheet = workbook.create_sheet("Report_Index")
    index_sheet.append(["Sheet", "Purpose"])
    for cell in index_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    meta_sheet = workbook.create_sheet("Artifact_Metadata")
    meta_sheet.append(["Key", "Value"])
    for cell in meta_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for key, value in (metadata or {}).items():
        meta_sheet.append([key, _json_safe(value)])

    tables = dict(tables or {})
    if feature_importance is not None:
        tables.setdefault("Feature_Importance", feature_importance)

    purposes = {
        "Dev_Metrics": "Development-set performance",
        "OOT_Metrics": "Out-of-time performance",
        "Dev_Score_Bins": "Development score ranking and lift",
        "OOT_Score_Bins": "OOT score ranking and lift",
        "IV_Summary": "Information value summary",
        "Selection_Report": "Feature selection audit",
        "Variable_Audit": "Feature lifecycle audit",
        "Binning_Summary": "Bin-level WOE/IV diagnostics",
        "WOE_Detail": "Detailed WOE mapping",
        "Score_PSI": "Score distribution stability",
        "Stability_Summary": "Stability quality gate",
        "Segment_Summary": "Segment-level performance",
        "Temporal_Stability": "Cross-period stability",
        "Benchmark_Performance": "Benchmark comparison",
        "Model_Estimation": "Model coefficients or estimator metadata",
        "Feature_Importance": "Model feature importance",
    }

    existing_names = {sheet.title for sheet in workbook.worksheets}
    for sheet_name, table in tables.items():
        safe_name = str(sheet_name)[:31] or "Table"
        if safe_name in existing_names:
            safe_name = f"{safe_name[:27]}_tbl"
        existing_names.add(safe_name)
        sheet = workbook.create_sheet(safe_name)
        _append_table(sheet, table, header_font=header_font)
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.fill = header_fill
        index_sheet.append([safe_name, purposes.get(str(sheet_name), "Audit table")])

    index_sheet.append(["Overview_Performance", "Primary metrics"])
    index_sheet.append(["Artifact_Metadata", "Reproducibility and deployment metadata"])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            width = min(
                48,
                max(len(str(cell.value or "")) for cell in column_cells) + 2,
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions

    workbook.save(report_path)
    logger.info(f"Model report saved to {report_path}")
    return report_path
