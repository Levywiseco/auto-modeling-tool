"""Excel audit report contract tests."""

from pathlib import Path

import polars as pl

from src.reports.excel import write_model_report


def test_model_report_contains_stable_core_sheets(tmp_path: Path):
    report_path = write_model_report(
        tmp_path,
        {"auc_roc": 0.8, "ks_statistic": 0.4},
        feature_importance=pl.DataFrame(
            {"Feature": ["x"], "Importance": [0.9]}
        ),
        metadata={"artifact_version": "1.0"},
        filename="Model_Report_1.xlsx",
    )
    assert report_path == tmp_path / "Model_Report_1.xlsx"
    from openpyxl import load_workbook

    workbook = load_workbook(report_path, read_only=True)
    assert {
        "Overview_Performance",
        "Feature_Importance",
        "Artifact_Metadata",
    }.issubset(set(workbook.sheetnames))
