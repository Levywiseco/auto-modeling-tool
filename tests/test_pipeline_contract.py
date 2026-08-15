"""Regression tests for the Dev/OOT and scoring contracts."""

import joblib
import numpy as np
import polars as pl
import pytest
from sklearn.linear_model import LogisticRegression

from auto_modeling_tool.binning.woe_binning import WoeBinner
from auto_modeling_tool.data.preprocess import DataPreprocessor
from auto_modeling_tool.data.split import split_dev_oot
from auto_modeling_tool.modeling.scorecard import (
    ScorecardBuilder,
    probability_to_credit_score,
)


def test_explicit_sample_dev_oot_split():
    data = pl.DataFrame(
        {
            "feature": list(range(8)),
            "target": [0, 1, 0, 1, 0, 1, 0, 1],
            "sample": ["dev", "dev", "dev", "dev", "oot", "oot", "oot", "oot"],
        }
    )
    result = split_dev_oot(data, "target", sample_column="sample")
    assert result.strategy == "sample_column"
    assert len(result.dev) == 4
    assert len(result.oot) == 4


def test_date_dev_oot_split():
    data = pl.DataFrame(
        {
            "feature": [1, 2, 3, 4],
            "target": [0, 1, 0, 1],
            "month": ["2024-01", "2024-02", "2024-03", "2024-04"],
        }
    )
    result = split_dev_oot(
        data,
        "target",
        date_column="month",
        oot_start="2024-03",
    )
    assert result.strategy == "date"
    assert len(result.dev) == 2
    assert len(result.oot) == 2


def test_preprocessor_uses_dev_statistics_only():
    dev = pl.DataFrame({"feature": [0.0, 1.0]})
    oot = pl.DataFrame({"feature": [100.0, 101.0]})
    preprocessor = DataPreprocessor(
        clean_strategy="median",
        normalize_method="zscore",
    )
    preprocessor.fit(dev)
    transformed = preprocessor.transform(oot)
    assert transformed["feature"].mean() > 100


def test_scorecard_uses_bin_index_for_points():
    X = pl.DataFrame({"feature": np.linspace(0, 1, 20)})
    y = pl.Series("target", [0, 1] * 10)
    binner = WoeBinner(n_bins=2, min_samples_bin=1)
    woe = binner.fit_transform(X, y, return_type="woe").select(["feature_bin"])
    model = LogisticRegression(max_iter=1000).fit(woe.to_numpy(), y.to_numpy())
    scorecard = ScorecardBuilder(
        base_score=600,
        PDO=20,
        target_odds=20,
    ).fit(model, binner, feature_names=["feature_bin"])
    table = scorecard.get_scorecard_table()
    expected = sorted(binner.bin_woes_["feature"].values())
    actual = sorted(table["WOE"].to_list())
    np.testing.assert_allclose(actual, expected, rtol=1e-6, atol=1e-6)
    scores = scorecard.score(X)
    assert len(scores) == len(X)


def test_encoding_safe_stream_handler_replaces_unrepresentable_text():
    import logging

    from auto_modeling_tool.core.logger import EncodingSafeStreamHandler

    class GbkStream:
        encoding = "gbk"

        def __init__(self):
            self.values = []

        def write(self, value):
            value.encode(self.encoding)
            self.values.append(value)

        def flush(self):
            pass

    stream = GbkStream()
    handler = EncodingSafeStreamHandler(stream)
    handler.setFormatter(logging.Formatter("%(message)s"))
    handler.emit(logging.LogRecord("test", logging.INFO, "<test>", 1, "✅ safe", (), None))

    assert "".join(stream.values) == "? safe\n"


def test_weighted_xgboost_pipeline_and_release_gate(tmp_path):
    from auto_modeling_tool.evaluation.quality_gate import validate_release
    from auto_modeling_tool.pipelines.auto_pipeline import AutoPipeline

    rng = np.random.default_rng(7)
    n_rows = 120
    x1 = rng.normal(size=n_rows)
    x2 = rng.normal(size=n_rows)
    target = (x1 + 0.35 * x2 > 0).astype(int)
    frame = pl.DataFrame({
        "x1": x1,
        "x2": x2,
        "target": target,
        "weight": np.where(target == 1, 2.0, 1.0),
        "sample": ["dev"] * 90 + ["oot"] * 30,
    })

    pipeline = AutoPipeline(
        target_col="target",
        model_type="xgboost",
        model_params={
            "n_estimators": 20,
            "max_depth": 2,
            "learning_rate": 0.1,
        },
        n_bins=4,
        n_features=2,
        early_stopping_eval="dev_holdout",
        early_stopping_rounds=3,
    )
    pipeline.fit(
        frame,
        sample_col="sample",
        weight_col="weight",
        use_sample_weight=True,
        min_samples_bin=5,
    )
    metrics = pipeline.evaluate()
    assert "auc_roc" in metrics
    # Dev/OOT score PSI must be a metric, not just a report table: the release
    # gate reads it from the artifact, so --max-psi is inert without it.
    assert "score_psi" in metrics

    output_dir = pipeline.save(tmp_path)
    report_path = next(tmp_path.glob("Model_Report_*.xlsx"))
    assert (output_dir / "scoring_artifact.pkl").exists()

    artifact = joblib.load(output_dir / "scoring_artifact.pkl")
    assert artifact["metadata"]["metrics"]["score_psi"] == pytest.approx(
        metrics["score_psi"]
    )

    # A ceiling this loose always clears; the point is that the gate reads a
    # real number. 120 rows give a naturally large Dev/OOT PSI.
    result = validate_release(
        output_dir,
        report_path=report_path,
        min_auc=0.5,
        max_psi=10.0,
    )
    assert result.passed
    psi_check = next(c for c in result.checks if c.name == "max_psi")
    assert "psi=None" not in psi_check.detail

    # An unreachable PSI ceiling must fail the gate rather than pass silently.
    strict = validate_release(output_dir, report_path=report_path, max_psi=-1.0)
    assert not strict.passed
    assert any(
        check.name == "max_psi" and not check.passed for check in strict.checks
    )


def test_probability_to_credit_score_contract():
    scores = probability_to_credit_score(
        np.array([0.5, 0.1, 0.99]),
        base_score=500,
        pdo=50,
        min_score=300,
        max_score=900,
    )
    np.testing.assert_allclose(scores[0], 500.0, atol=1e-8)
    assert scores[1] > scores[0] > scores[2]
    assert np.all((scores >= 300) & (scores <= 900))


def test_export_excel_can_be_disabled(tmp_path):
    from auto_modeling_tool.pipelines.auto_pipeline import AutoPipeline

    frame = pl.DataFrame(
        {
            "feature": list(range(20)),
            "target": [int(index >= 10) for index in range(20)],
            "sample": ["dev"] * 14 + ["oot"] * 6,
        }
    )
    pipeline = AutoPipeline(
        target_col="target",
        n_bins=2,
        n_features=1,
        export_excel=False,
    )
    pipeline.fit(
        frame,
        sample_col="sample",
        export_excel=False,
        min_samples_bin=2,
    )
    pipeline.evaluate()
    pipeline.save(tmp_path)

    assert (tmp_path / "scoring_artifact.pkl").exists()
    assert not list(tmp_path.glob("Model_Report_*.xlsx"))



def test_evaluation_columns_are_not_trained_on(tmp_path):
    """segment/temporal/benchmark columns describe the model, they are not inputs.

    Training on a temporal column is time leakage — the model learns "March is
    bad" and is meaningless on an unseen month. Training on a benchmark column
    destroys the comparison and makes the external score a hard dependency of
    scoring. Both previously happened silently whenever those columns carried
    signal, inflating OOT AUC.
    """
    from auto_modeling_tool.main import run_modeling_pipeline

    rng = np.random.default_rng(5)
    n = 900
    x1 = rng.normal(size=n)
    month_idx = rng.integers(0, 3, n)
    # Both auxiliary columns are deliberately strong predictors.
    target = (x1 * 0.5 + month_idx * 1.5 + rng.normal(size=n) * 0.4 > 1.5).astype(int)
    csv = tmp_path / "d.csv"
    pl.DataFrame({
        "x1": x1,
        "apply_month": [f"2026-0{i + 1}" for i in month_idx],
        "bureau_score": target * 80 + rng.normal(600, 20, n),
        "channel": rng.choice(["app", "h5"], n),
        "target": target,
        "sample": ["dev"] * 700 + ["oot"] * 200,
    }).write_csv(csv)

    result = run_modeling_pipeline(
        str(csv), "target",
        output_dir=str(tmp_path / "out"),
        sample_col="sample", n_bins=5, min_samples_bin=20,
        temporal_col="apply_month",
        benchmark_cols=["bureau_score"],
        segment_cols=["channel"],
        archive_run=False,
    )

    aux = {"apply_month", "bureau_score", "channel"}
    assert not aux & set(result["selected_features"])
    assert not any(
        any(a in feature for a in aux) for feature in result["selected_features"]
    )

    # The scoring contract must not demand columns the model never uses.
    artifact = joblib.load(tmp_path / "out" / "scoring_artifact.pkl")
    assert not aux & set(artifact["feature_columns"])

    # Excluding them from training must not remove them from the report.
    from openpyxl import load_workbook

    report = next((tmp_path / "out").glob("Model_Report_*.xlsx"))
    sheets = set(load_workbook(report, read_only=True).sheetnames)
    assert {"Temporal_Stability", "Benchmark_Performance", "Segment_Summary"} <= sheets


def test_pipeline_fits_preprocessing_on_dev_only(tmp_path):
    """The pipeline — not just DataPreprocessor — must never see OOT when fitting.

    Mutation testing exposed this gap: changing the pipeline to fit the
    preprocessor on Dev+OOT left all 190 tests green. The existing preprocessor
    test only checked the transformer in isolation, so the leakage-safety the
    whole project is built around had no coverage at the pipeline level.
    """
    from auto_modeling_tool.pipelines.auto_pipeline import AutoPipeline

    rng = np.random.default_rng(17)
    n_dev, n_oot = 600, 200
    # OOT is shifted far from Dev. Statistics fitted on Dev alone must ignore it.
    dev_values = rng.normal(0.0, 1.0, n_dev)
    oot_values = rng.normal(50.0, 1.0, n_oot)
    values = np.concatenate([dev_values, oot_values])
    target = np.concatenate([
        (dev_values > 0).astype(int),
        (oot_values > 50).astype(int),
    ])
    frame = pl.DataFrame({
        "x": values,
        "noise": rng.normal(size=n_dev + n_oot),
        "target": target,
        "sample": ["dev"] * n_dev + ["oot"] * n_oot,
    })

    pipeline = AutoPipeline(target_col="target", n_bins=5, n_features=2)
    pipeline.fit(frame, sample_col="sample", min_samples_bin=20)

    stats = pipeline.preprocessor_.stats_["x"]
    dev_mean = float(dev_values.mean())
    pooled_mean = float(values.mean())

    # Dev mean is ~0; pooling in the shifted OOT block would drag it to ~12.5.
    assert abs(stats["mean"] - dev_mean) < 0.5, (
        f"preprocessor mean {stats['mean']:.3f} is not Dev's {dev_mean:.3f} — "
        f"OOT appears to have been included (pooled would be {pooled_mean:.3f})"
    )

    # Bin edges are fitted on Dev too, so no cut may sit out in OOT territory.
    assert max(c for c in pipeline.binner_.bin_cuts_["x"] if np.isfinite(c)) < 25.0


def test_credit_score_scale_follows_pdo():
    """PDO sets how many points double the odds; hardcoding it goes unnoticed.

    Mutation testing showed both the ScorecardBuilder factor and the standalone
    helper could ignore their configured PDO with every test still passing.
    """
    probabilities = np.array([0.2, 0.5, 0.8])

    narrow = probability_to_credit_score(
        probabilities, base_score=600, pdo=20, min_score=0, max_score=1000
    )
    wide = probability_to_credit_score(
        probabilities, base_score=600, pdo=80, min_score=0, max_score=1000
    )

    # A larger PDO stretches the same odds range over more points.
    assert (wide.max() - wide.min()) > (narrow.max() - narrow.min()) * 3

    # And the ratio tracks the PDO ratio directly.
    assert (wide.max() - wide.min()) / (narrow.max() - narrow.min()) == pytest.approx(
        4.0, rel=0.01
    )


def test_scorecard_builder_factor_follows_pdo():
    """factor_ is set during fit; hardcoding it there passed every test."""
    from auto_modeling_tool.modeling.scorecard import ScorecardBuilder

    X = pl.DataFrame({"feature": np.linspace(0, 1, 40)})
    y = pl.Series("target", [0, 1] * 20)
    binner = WoeBinner(n_bins=2, min_samples_bin=1)
    woe = binner.fit_transform(X, y, return_type="woe").select(["feature_bin"])
    model = LogisticRegression(max_iter=1000).fit(woe.to_numpy(), y.to_numpy())

    for pdo in (20, 50, 80):
        card = ScorecardBuilder(base_score=600, PDO=pdo, target_odds=20).fit(
            model, binner, feature_names=["feature_bin"]
        )
        assert card.factor_ == pytest.approx(pdo / np.log(2))
